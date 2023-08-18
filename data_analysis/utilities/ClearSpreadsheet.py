import sys
sys.path.append("/Users/zacharystarr/Desktop/TrafficMonitoring-usingYOLO")

from openpyxl import Workbook, load_workbook
import params

wb = load_workbook("{}/{}".format(params.SpreadsheetFolder, params.SpreadsheetName))
ws = wb.active
initialColumn = 2
currentColumn = 2
openColumn = False
while not openColumn:
    if str(ws.cell(1, currentColumn).value) == "None":
        openColumn = True
    else:
        currentColumn += 1
print("Deleting a total of {} Columns".format(currentColumn - 2))
ws.delete_cols(initialColumn, currentColumn)
wb.save("{}/{}".format(params.SpreadsheetFolder, params.SpreadsheetName))
print("Done!") 